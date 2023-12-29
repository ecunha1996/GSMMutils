# -*- coding: utf-8 -*-
"""
Created on Mon Jun 29 14:52:26 2020

@author: ecunha
"""
import copy
import os
import re
from os.path import join
from typing import Union

import cobra
import numpy as np
import pandas as pd
from cobra import flux_analysis, Model, Reaction, Metabolite
from cobra.flux_analysis import find_essential_genes
from cobra.io import write_sbml_model
from cobra.manipulation.delete import prune_unused_metabolites
from openpyxl import load_workbook
from pandas import DataFrame
from sympy import Add

from GSMMutils.experimental.Biomass import Biomass
from GSMMutils.experimental.BiomassComponent import BiomassComponent
from GSMMutils.io import write_simulation
from GSMMutils.utils.utils import update_st, get_precursors, normalize, convert_mmol_mol_to_g_molMM, \
    convert_mg_molMM_to_mmolM_gMM

import logging

logging.getLogger('cobra').setLevel(logging.CRITICAL)


class MyModel(Model):
    def __init__(self, file_name=None, biomass_reaction=None, directory=None, prune_metabolites=False):
        self.genes_pathways_map = None
        self.e_res_precursors = None
        self.bio_precursors_res = None
        self.precursors_reactions = None
        self.extra_compartment = None
        self.intra_compartment = None
        self.pathway_reactions_map = {}
        self.biomass_reaction = None
        if not directory:
            directory = os.getcwd()
        self.directory = directory
        self.file_name = file_name
        self.model_old = []
        self.model_first = None
        self._biomass_composition = None
        self.load_model(self.directory, self.file_name)
        if not biomass_reaction:
            biomass_reaction = self.search_biomass()
        if biomass_reaction:
            self.bio_reaction = self.model.reactions.get_by_id(biomass_reaction)
        if not self.biomass_reaction:
            self.biomass_reaction = self.bio_reaction.id
        potential_biomass_metabolite = [e for e in self.bio_reaction.products if 'biomass' in e.id.lower()]
        if potential_biomass_metabolite:
            self.biomass_metabolite = potential_biomass_metabolite[0]
        else:
            self.biomass_metabolite = Metabolite("e_Biomass__cytop")
        # self.set_compartments()
        self.reactions_pathways_map = None
        self.bio_precursors = None
        self.pre_precursors = None
        # self.parse_reactions_versions()
        if prune_metabolites:
            self.model = prune_unused_metabolites(self.model)[0]
        super().__init__(self.model)
        if self.bio_reaction:
            self.objective = self.bio_reaction.id
        self.biomass_components = {}
        # self.parse_genes()
        self.model_first = self.model
        self.get_pathway_reactions_map()
        self.reaction_ids = [reaction.id for reaction in self.reactions]
        print("Reactions:", len(self.model.reactions))
        print("Metabolites:", len(self.model.metabolites))
        print("Genes:", len(self.model.genes))
        print("Model loaded")

    @property
    def biomass_composition(self):
        return self._biomass_composition

    @biomass_composition.setter
    def biomass_composition(self, biomass_composition=None):
        if type(biomass_composition) == str or not biomass_composition:
            self._biomass_composition = {}
            if not self.biomass_components:
                self.infer_biomass_from_model()
            biomass_composition = {}
            for children in self.biomass_components[self.biomass_metabolite.id].children:
                biomass_composition[children.id] = children.stoichiometry
            self._biomass_composition = biomass_composition
        elif type(biomass_composition) == Biomass:
            self._biomass_composition = biomass_composition
        else:
            raise TypeError("Biomass composition must be a string or a Biomass object")

    def load_model(self, directory, file_name):
        """
        This function loads the model. Returns a COBRAApy object, the model
        It also creates a copy of the model to be used in the first simulation
        :param directory:
        :param file_name:
        :return:
        """

        print("Loading")
        print("")

        os.chdir(directory)

        self.model = cobra.io.read_sbml_model(join(directory, file_name))
        if not self.model.exchanges:
            for reaction in self.model.reactions:
                if "EX_" in reaction.id:
                    met = reaction.products[0]
                    reaction.add_metabolites({met: -1})
        # print("deleting b")
        self.delete_b_metabolites()
        return self.model

    def copy(self) -> 'MyModel':
        return copy.deepcopy(self)

    def set_compartments(self):
        compartments = self.model.compartments.keys()
        for key in compartments:
            if "outside" in self.model.compartments[key]:
                self.extra_compartment = key
            if "inside" in self.model.compartments[key]:
                self.intra_compartment = key

    def search_biomass(self):
        for reaction in self.model.reactions:
            if "biomass" in reaction.id.lower():
                return reaction

    def delete_b_metabolites(self):
        for metabolite in self.model.metabolites:
            if '_b' == metabolite.id[-2:] and 'EX_' in metabolite.id:
                self.model.remove_metabolites([metabolite])
                print(metabolite.id + "was removed")

    def get_reaction(self, reaction: Union[str, Reaction]):
        try:
            if type(reaction) == str:
                return self.reactions.get_by_id(reaction)
            elif type(reaction) == Reaction:
                return reaction
        except KeyError:
            print(reaction + " not found")
        except Exception as e:
            print(e)

    def get_metabolite(self, metabolite):
        return self.model.metabolites.get_by_id(metabolite)

    def get_reactants(self, reaction):
        return self.get_reaction(reaction).reactants

    def get_products(self, reaction):
        return self.get_reaction(reaction).products

    def get_bio_reaction(self):
        return self.bio_reaction

    def get_exchanges(self):
        return self.model.exchanges

    def get_bio_precursors(self):
        """
        This function returns the precursors of the biomass reaction, i.e., the reactants of the biomass reaction.
        Returns
        -------

        """

        self.bio_precursors = []
        reactants = self.get_reactants(self.bio_reaction.id)
        for reactant in reactants:
            self.bio_precursors.append(reactant)

        return self.bio_precursors

    def get_pre_precursors(self):
        """
        This function returns the precursors of the biomass precursors, i.e., the reactants of the reactions that
        produce the biomass precursors.
        Returns
        -------

        """
        try:
            self.pre_precursors = {}
            self.precursors_reactions = {}
            if self.bio_precursors is None:
                self.get_bio_precursors()
            for precursor in self.bio_precursors:
                reactions = precursor.reactions
                if len(reactions) > 2:
                    print("ATP or other metabolite. Without unique precursor")
                    self.pre_precursors[precursor.id] = []
                else:
                    reaction = Reaction()
                    for r in reactions:
                        if r != self.bio_reaction:
                            reaction = r
                    if reaction.id:
                        self.pre_precursors[precursor.id] = self.get_reactants(reaction.id)
                        self.precursors_reactions[precursor.name] = (reaction.id, precursor.id)
        except Exception as e:
            print(e)
        return self.pre_precursors

    def set_stoichiometry(self, reaction, metabolite, stoichiometry):
        self.get_reaction(reaction).add_metabolites({
            self.get_metabolite(metabolite): -self.get_reaction(reaction).metabolites[self.get_metabolite(metabolite)]})
        self.get_reaction(reaction).add_metabolites({self.get_metabolite(metabolite): stoichiometry})

    def maximize(self, value=True, pfba=True):
        """
        This function maximizes the biomass reaction. It can return the biomass flux or the whole solution
        Parameters
        ----------
        value: bool
        pfba: bool

        Returns
        -------
        """
        try:
            if value:
                if pfba:
                    return flux_analysis.pfba(self.model)[self.biomass_reaction]
                else:
                    return self.model.optimize().objective_value
            else:
                if pfba:
                    return flux_analysis.pfba(self.model)
                else:
                    return self.model.optimize()
        except Exception as e:
            print(e)
            return 0

    def summary(self, solution=None, pfba=True, **kwargs):
        if solution:
            return self.model.summary(solution, **kwargs)
        if not pfba or kwargs:
            return self.model.summary(**kwargs)
        else:
            print("Running pFBA")
            try:
                pfba_sol = cobra.flux_analysis.pfba(self)
                return self.model.summary(pfba_sol, **kwargs)
            except Exception as e:
                print(e)
                return 0

    def create_reaction(self, reaction: Union[cobra.Reaction, str], metabolites: dict, **kwargs):
        """
        This function creates a reaction in the model.
        Parameters
        ----------
        reaction: Union[cobra.Reaction, str]
            Reaction to be created
        metabolites: dict
            Dictionary with the metabolites and their stoichiometry
        Returns
        -------
        cobra.Reaction
            The created reaction
        """
        if not isinstance(reaction, cobra.Reaction):
            new_reaction = Reaction(id=reaction, **kwargs)
            new_reaction.add_metabolites(metabolites)
            self.add_reactions([new_reaction])
        else:
            self.add_reactions([reaction])
        return self.get_reaction(reaction)

    def create_metabolite(self, metabolite, is_boundary_metabolite=True, **kwargs):
        if is_boundary_metabolite:
            compartment = self.extra_compartment
            metabolite_id = "Ex_" + metabolite
        else:
            compartment = self.intra_compartment
            metabolite_id = "In_" + metabolite

        self.add_metabolites([cobra.Metabolite(metabolite_id, compartment=compartment, **kwargs)])
        return self.get_metabolite(metabolite_id)

    def parse_reactions_versions(self):
        for reaction in self.model.reactions:
            for i in range(10):
                if "_V" + str(i) in reaction.id:
                    try:
                        reaction.id = reaction.id.replace("_V1", "")
                        reaction.id = reaction.id.replace("_V2", "")
                        reaction.id = reaction.id.replace("_V3", "")
                        reaction.id = reaction.id.replace("_V4", "")
                        reaction.id = reaction.id.replace("_V5", "")
                        reaction.id = reaction.id.replace("_V6", "")
                    except Exception as e:
                        print(e)

    def create_exchange(self, metabolite_id, bounds=(-10000, 10000)):

        reaction_name = "EX_" + metabolite_id
        self.create_reaction(reaction_name, {self.get_metabolite(metabolite_id): -1})
        self.get_reaction(reaction_name).bounds = bounds
        return self.get_reaction(reaction_name)

    def create_transport(self, ex_metabolite, in_metabolite, uptake=True, bounds=(-10000, 10000)):

        if uptake:
            stoichiometry = 1
        else:
            stoichiometry = -1

        reaction_name = "Tr_" + ex_metabolite + "_" + in_metabolite
        self.create_reaction(reaction_name,
                             {self.get_metabolite(ex_metabolite): -stoichiometry,
                              self.get_metabolite(in_metabolite): stoichiometry})
        self.get_reaction(reaction_name).bounds = bounds
        return self.get_reaction(reaction_name)

    def create_demand(self, metabolite_id):
        """
        This function creates a demand reaction
        Parameters
        ----------
        metabolite_id: str
            Metabolite id

        Returns
        -------
        cobra.Reaction
            Created demand reaction
        """
        reaction_name = "DM_" + metabolite_id
        self.create_reaction(reaction_name, {self.get_metabolite(metabolite_id): -1})
        self.get_reaction(reaction_name).bounds = (0, 10000)
        return self.get_reaction(reaction_name)

    def create_sink(self, metabolite_id, bounds=(-10000, 10000)):
        """
        This function creates a sink reaction
        Parameters
        ----------
        metabolite_id: str
            Metabolite id
        bounds: tuple
            Bounds of the reaction
        Returns
        -------
        cobra.Reaction
            The created reaction
        """
        reaction_name = "Sk_" + metabolite_id
        self.create_reaction(reaction_name, {self.get_metabolite(metabolite_id): 1})
        self.get_reaction(reaction_name).bounds = bounds
        return self.get_reaction(reaction_name)

    def create_exchange_transport(self, ex_metabolite, in_metabolite, uptake=True, bounds=(-10000, 10000)):
        """
        This function creates an exchange and a transport reaction
        Parameters
        ----------
        ex_metabolite: str
            Metabolite id
        in_metabolite: str
            Metabolite id
        uptake: bool
            If the transport is uptake or secretion
        bounds: tuple
            Bounds of the reaction

        Returns
        -------
        tuple
            The created reactions (exchange, transport)
        """
        drain = self.create_exchange(ex_metabolite, bounds)
        transport = self.create_transport(ex_metabolite, in_metabolite, uptake, bounds)
        return drain, transport

    def save(self):
        if len(self.model_old) > 5:
            self.model_old.pop()
        self.model_old.insert(0, self.model.copy())

    def undo(self):
        previous_model = self.model_old.pop(0)
        self.model = previous_model.copy()

    def reset(self):
        self.model = self.model_first

    def test_bio_precursors(self):

        """This function tests any precursor of the biomass. 
        Reactants in the biomass equation"""

        if self.bio_precursors is None:
            self.get_bio_precursors()

        self.save()

        bio_precursors_res = {"Flux": []}

        meta_val = []

        for precursor in self.bio_precursors:
            self.save()
            reaction = self.create_demand(precursor.id)
            self.objective = reaction.id
            bio_precursors_res["Flux"].append(self.maximize())
            meta_val.append(precursor.name)
            self.undo()

        self.bio_precursors_res = DataFrame(bio_precursors_res, index=meta_val)

        return self.bio_precursors_res

    def evaluate_precursor(self, current_precursor):
        self.save()
        reaction = self.create_demand(current_precursor.id)
        self.objective = reaction.id
        val = self.maximize()
        self.undo()
        return val

    def test_e_precursors(self):
        """
        This function tests any precursor of the biomass.
        Returns
        -------

        """

        if self.bio_precursors is None:
            self.get_bio_precursors()
        if self.pre_precursors is None:
            self.get_pre_precursors()

        self.save()

        e_precursors_res = {"Flux": []}
        meta_val = []

        for precursor in self.bio_precursors:
            e_precursors_res["Flux"].append(self.evaluate_precursor(precursor))
            meta_val.append(precursor.name)
            for pre_precursor in self.pre_precursors[precursor.id]:
                e_precursors_res["Flux"].append(self.evaluate_precursor(pre_precursor))
                meta_val.append(pre_precursor.name)

        self.e_res_precursors = DataFrame(e_precursors_res, index=meta_val)

        return self.e_res_precursors

    def test_e_reaction(self, e_metabolite):
        """
        This function tests any precursor of the biomass (known as e-metabolites), which are reactants in the biomass
        Parameters
        ----------
        e_metabolite: str
            Metabolite id

        Returns
        -------

        """

        def evaluate(current_precursor):
            self.save()
            reaction = self.create_demand(current_precursor.id)
            self.objective = reaction.id
            val = self.maximize()
            e_precursors_res["Flux"].append(val)
            meta_val.append(current_precursor.name)
            self.undo()

        if self.pre_precursors is None:
            self.get_pre_precursors()

        self.save()

        e_precursors_res = {"Flux": []}
        meta_val = []

        for pre_precursor in self.pre_precursors[e_metabolite]:
            evaluate(pre_precursor)

        res = DataFrame(e_precursors_res, index=meta_val)

        return res

    def test_reaction(self, reaction):
        """
        This function tests any precursor of the biomass (known as e-metabolites), which are reactants in the biomass
        Parameters
        ----------
        reaction: str
            Reaction id
        Returns
        -------

        """
        if reaction.startswith("R_"):
            reaction = reaction.replace("R_", "")

        pattern = re.compile('[^a-zA-Z0-9_]')

        reaction = re.sub(pattern, '_', reaction)

        old_objective = self.objective
        precursors = self.get_reactants(reaction)

        e_precursors_res = {"Flux": [], "ReactantOrProduct": []}
        meta_val = []
        demands_ids = [demand.id for demand in self.demands]
        for precursor in precursors:
            if "DM_" + precursor.id not in demands_ids:
                try:
                    temp_reaction = self.create_demand(precursor.id)
                except Exception as e:
                    print(e)
                    continue
            else:
                temp_reaction = self.get_reaction("DM_" + precursor.id)
            self.objective = temp_reaction.id
            val_2 = self.model.optimize().objective_value
            e_precursors_res["Flux"].append(val_2)
            e_precursors_res["ReactantOrProduct"].append("Reactant")
            meta_val.append(precursor.name)
            self.remove_reactions(temp_reaction.id)
        sinks_ids = [sink.id for sink in self.sinks]
        for product in self.get_products(reaction):
            if "Sk_" + product.id not in sinks_ids:
                try:
                    temp_reaction = self.create_sink(product.id)
                except Exception as e:
                    print(e)
                    continue
            else:
                temp_reaction = self.get_reaction("Sk_" + product.id)
            self.objective = temp_reaction.id
            val_2 = self.model.optimize().objective_value
            e_precursors_res["Flux"].append(val_2)
            e_precursors_res["ReactantOrProduct"].append("Product")
            meta_val.append(product.name)
            self.remove_reactions(temp_reaction.id)
        res = DataFrame(e_precursors_res, index=meta_val)
        self.objective = old_objective
        return res

    def create_test_drains_to_reaction(self, reaction, reactants=True, products=False, uptake=True,
                                       bounds=(-10000, 10000)):
        """

        Parameters
        ----------
        reaction: str
            id of the reaction
        reactants: bool
            if True, drains for all reactants will be created
        products: bool
            if True, drains for all products will be created
        uptake: bool
            if True, the drain will be an uptake reaction
        bounds: tuple
            bounds of the drain reaction

        Returns
        -------

        """
        self.save()

        if reactants:
            reactant = self.get_reactants(reaction)
            for r in reactant:
                ex_m = self.create_metabolite(r.id)
                self.create_exchange_transport(ex_m.id, r.id, uptake, bounds)
            print("Drains for all reactants created")

        elif reactants and products:
            reactant = self.get_reactants(reaction)
            product = self.get_products(reaction)
            both = reactant + product
            for r in both:
                ex_m = self.create_metabolite(r.id)
                self.create_exchange_transport(ex_m.id, r.id, uptake, bounds)
            print("Drains for all reactants and products created")

        else:
            product = self.get_products(reaction)
            for r in product:
                ex_m = self.create_metabolite(r.id)
                self.create_exchange_transport(ex_m.id, r.id, uptake, bounds)
            print("Drains for all products created")

    def create_trnas_reactions(self, protein_id="e-Protein"):
        """
        Function to create tRNA synthesis reactions for protein
        Parameters
        ----------
        protein_id: str
            id of the protein

        Returns
        -------

        """
        if self.pre_precursors is None:
            self.get_pre_precursors()

        self.save()

        trnas = self.get_products(self.precursors_reactions[protein_id][0])

        for trna in trnas:
            if "H2O" not in trna.name and "e-Protein" not in trna.name:
                self.create_sink(trna.id)

    def create_reporting_file(self, spreadsheet_name):
        """
        Function to create a spreadsheet with the results of the test
        Parameters
        ----------
        spreadsheet_name: str
            name of the spreadsheet

        Returns
        -------

        """
        res = self.test_e_precursors()

        res["Alterations into the model"] = " "

        workbook = pd.ExcelWriter(spreadsheet_name)
        res.to_excel(workbook)
        workbook.save()

        print(spreadsheet_name, "created")

    def create_fluxes(self):
        """
        This function creates a DataFrame with the fluxes of the model
        Returns
        -------
        DataFrame
            Pandas DataFrame with the fluxes of the model
        """
        self.objective = self.bio_reaction.id

        solution = self.optimize()

        return solution.fluxes

    def merge_fluxes(self, reactions_file_name, version):
        """
        This function merges the fluxes of the model with the reaction information
        Parameters
        ----------
        reactions_file_name: str
            name of the Excel file with the reaction information
        version: str
            version of the model

        Returns
        -------
        DataFrame
            Pandas DataFrame with the fluxes of the model and the reaction information
        """
        fluxes = self.create_fluxes()

        merlin_file = pd.read_excel(reactions_file_name)

        fluxes_ids = list(fluxes.index)

        fluxes_names = [self.get_reaction(reaction).name for reaction in fluxes_ids]

        fluxes_names_2 = [name.split("__")[0] for name in fluxes_names]

        fluxes_values = [fluxes[val] for val in fluxes_ids]

        flux = list(zip(fluxes_ids, fluxes_names_2, fluxes_values))

        res = []

        res_model_id = []

        for f in flux:
            pathway = merlin_file["Pathway Name"][merlin_file["Reaction Name"] == f[1]].values
            equation = merlin_file["Equation"][merlin_file["Reaction Name"] == f[1]].values
            for path in pathway:
                res.append((f[1], path, equation[0], f[2]))
                res_model_id.append(f[0])

        df = DataFrame(data=res, index=res_model_id,
                       columns=["Reaction ID", "Pathway", "Equation", "Flux " + version])

        return df

    def create_fluxes_spreadsheet(self, reactions_file_name, spreadsheet_name, version):
        """
        This function creates a spreadsheet with the fluxes of the model and the reaction information
        Parameters
        ----------
        reactions_file_name: str
            name of the Excel file with the reaction information
        spreadsheet_name: str
            name of the spreadsheet
        version: str
            version of the model

        Returns
        -------

        """

        df = self.merge_fluxes(reactions_file_name, version)

        workbook = pd.ExcelWriter(spreadsheet_name)
        df.to_excel(workbook)
        workbook.save()

        print(spreadsheet_name, "created")

    def add_medium(self, medium_file_name, medium_sheet_name):
        """
        This function adds the medium to the model from an Excel file
        Parameters
        ----------
        medium_file_name: str
            name of the Excel file with the medium
        medium_sheet_name: str
            name of the sheet with the medium

        Returns
        -------

        """
        self.save()

        file = pd.read_excel(medium_file_name, medium_sheet_name, converters={"Model ID": str}, engine="openpyxl")

        for reaction in file["Reaction ID"]:
            try:
                self.reactions.get_by_id(reaction).lower_bound = float(file["LB"][file["Reaction ID"] == reaction].values[0])
                self.reactions.get_by_id(reaction).upper_bound = float(file["UB"][file["Reaction ID"] == reaction].values[0])
            except Exception as e:
                print(e)

    def write(self, filename):
        """
        This function writes the model in SBML format
        Parameters
        ----------
        filename: str
            name of the file to write the model

        Returns
        -------

        """
        write_sbml_model(self, filename)

    def update_fluxes_spreadsheet(self, file_name, sheet_name, column_to_write, version, medium):
        """
        This function updates the flux spreadsheet with the fluxes of the model
        Parameters
        ----------
        file_name: str
            name of the spreadsheet
        sheet_name: str
            name of the sheet
        column_to_write: str
            column to write the fluxes
        version: str
            version of the model
        medium: str
            name of the medium

        Returns
        -------

        """
        wb = load_workbook(file_name)

        ws = wb[sheet_name]

        model_ids = ws["A"]

        fluxes = self.create_fluxes()

        for model_id in model_ids:

            i = model_id.row

            i_2 = column_to_write + str(i)

            if i != 1:
                ws[i_2].value = fluxes[model_id.value]
            else:
                ws[i_2].value = "Flux " + version + " / " + medium

        wb.save(file_name)

    def get_metabolite_by_name(self, name, compartment="C00002"):
        """
        This function returns a metabolite by its name
        Parameters
        ----------
        name: str
            The name of the metabolite
        compartment: str
            The compartment of the metabolite

        Returns
        -------
        Metabolite
            The metabolite with the name and compartment
        """
        for met in self.model.metabolites:
            if name == met.name and met.compartment == compartment:
                return met
        print(name + "\t" * 2 + "Not found")
        return None

    def apply_env_conditions_from_excel(self, conditions_file_name, conditions_sheet_name):
        """
        This function applies the environmental conditions from an Excel file
        Parameters
        ----------
        conditions_file_name: str
            Name of the Excel file with the environmental conditions
        conditions_sheet_name: str
            Name of the sheet with the environmental conditions

        Returns
        -------

        """
        file = pd.read_excel(conditions_file_name, conditions_sheet_name)
        for exchange in self.exchanges:
            try:
                exchange.lower_bound = float(-file["Uptake Rate"][file["Exchange"] == exchange.id].values[0])
            except Exception as e:
                print("Error applying environmental conditions")
                print(e)

    def apply_env_conditions_from_dict(self, data, metabolites=None, aliases=None):
        try:
            for metabolite in metabolites:
                if metabolite in data.keys():
                    if type(data[metabolite]) != tuple:
                        data[metabolite] = (-data[metabolite], 1000)
                    if aliases and metabolite in aliases.keys():
                        metabolite_in_model = aliases[metabolite]
                    else:
                        metabolite_in_model = metabolite
                    self.reactions.get_by_id(f"EX_{metabolite_in_model}__dra").bounds = data[metabolite]
        except Exception as e:
            print("Error applying environmental conditions")
            print(e)

    def minimal_medium(self, conditions_file_name, conditions_sheet_name, output_file_name, minimal_growth):
        """
        This function calculates the minimal medium for a given growth rate
        Parameters
        ----------
        conditions_file_name: str
        conditions_sheet_name: str
        output_file_name: str
        minimal_growth: float

        Returns
        -------

        """
        writer = pd.ExcelWriter(os.path.join(self.directory, output_file_name), engine='xlsxwriter')

        file = pd.read_excel(conditions_file_name, conditions_sheet_name)

        compounds = file["Exchange"].unique()
        exchanges = self.get_exchanges()

        for exchange in exchanges:
            self.get_reaction(exchange.id).lower_bound = float(
                -file["Uptake Rate"][file["Exchange"] == exchange.id].values[0])

        exchanges_2 = [exchange for exchange in exchanges]
        output = []
        output_ids = []
        bounds = {}

        for exchange in exchanges:
            print()
            if file["Uptake Rate"][file["Exchange"] == exchange.id].values[0] > 0.01:
                print()
                if exchange.id in compounds:
                    print()
                    bounds[exchange.id] = exchange.lower_bound
                    self.get_reaction(exchange.id).lower_bound = 0.0
                    print()
                    try:
                        biomass_value = self.maximize(pfba=True)
                    except Exception as e:
                        biomass_value = 0
                        print(e)

                    self.get_reaction(exchange.id).lower_bound = bounds[exchange.id]

                    if biomass_value <= minimal_growth:
                        output.append((exchange.id, biomass_value, "level 1"))
                        output_ids.append(exchange.id)
                        exchanges_2.remove(exchange)
                else:
                    exchanges_2.remove(exchange)

            else:
                exchanges_2.remove(exchange)
        print()
        for exchange in exchanges_2:
            self.get_reaction(exchange.id).lower_bound = 0.0

        def get_combinations(compounds_to_combine, current_exchanges, current_level, current_extra_sources,
                             minimal_growth_rate):
            current_exchange = current_exchanges[0]
            if current_exchange.id in compounds_to_combine:

                self.get_reaction(current_exchange.id).lower_bound = bounds[current_exchange.id]
                bio_val = self.maximize(pfba=True)
                if bio_val > minimal_growth_rate:
                    current_extra_sources.append((current_exchange, current_level + current_exchange.id))
                    for extra_source in current_extra_sources:
                        self.get_reaction(extra_source[0].id).lower_bound = 0.0
                    output.append((current_extra_sources[0][0].name, bio_val, current_extra_sources[-1][1]))
                    output_ids.append(current_extra_sources[0][0].id)
                else:
                    current_level = current_level + current_exchange.id + " ; "
                    current_extra_sources.append((current_exchange, current_level))
                    ex = current_exchanges.pop(0)
                    current_exchanges.append(ex)
                    get_combinations(compounds_to_combine, current_exchanges, current_level, current_extra_sources,
                                     minimal_growth_rate)

        feasible = self.maximize(pfba=True)

        if feasible and self.maximize(pfba=True) > minimal_growth:

            df = DataFrame(data=output, index=output_ids, columns=["Name", "Biomass", "Level"])
            df.to_excel(writer, "Output")
            writer.save()
            writer.close()

        else:
            n_ex = len(exchanges_2)
            exchanges_3 = [exchange for exchange in exchanges_2]

            for _ in range(n_ex):
                extra_sources = []
                level = "level 2 ;"
                get_combinations(compounds, exchanges_2, level, extra_sources, minimal_growth)

                reaction = exchanges_3.pop(0)
                exchanges_3.append(reaction)
                exchanges_2 = [exchange for exchange in exchanges_3]

            df = DataFrame(data=output, index=output_ids, columns=["Name", "Biomass", "Level"])
            df.to_excel(writer, "Output")
            writer.save()
            writer.close()

    def minimize_uptake_sum(self, substrates=None, to_minimize=None, to_maximize=None):
        if substrates is None:
            substrates = [ex.id for ex in self.exchanges]
        qp_expression = [arg for substrate in substrates for arg in self.get_reaction(substrate).flux_expression.args]

        if to_minimize:
            qp_expression = [arg for arg in self.get_reaction(to_minimize).flux_expression.args]
            for arg in qp_expression:
                temp = list(arg.args)
                temp[0] = -1
                arg.args = temp
            qp_expression += [arg for arg in self.get_reaction(to_maximize).flux_expression.args]

        qp_expression = Add(*qp_expression)

        qp_objective = self.problem.Objective(qp_expression, direction='max')

        self.objective = qp_objective

    def gene_essentially(self, conditions_file_name, conditions_sheet_name):

        file = pd.read_excel(conditions_file_name, conditions_sheet_name)

        for exchange in self.exchanges:
            exchange.lower_bound = float(-file["Uptake Rate"][file["Exchange"] == exchange.id].values[0])

        print(self.maximize(pfba=True))

        deletion_results = find_essential_genes(self, processes=1)

        return deletion_results

    def connectivity(self, output_file_name, extracellular_compartment, cytosol_compartment, periplasm=False):

        # open output file
        writer = pd.ExcelWriter(os.path.join(self.directory, output_file_name), engine='xlsxwriter')

        col = ["Name", "Outside", "Inside"]

        data = {}

        metabolites = self.metabolites

        for met in metabolites:

            met_id = str(met.id)

            if met.compartment == extracellular_compartment:
                if met_id in data:
                    data[met_id][1] = len(met.reactions)
                else:
                    data[met_id] = [met.name, len(met.reactions), 0]

            if met.compartment == cytosol_compartment:
                if met_id in data:
                    data[met_id][2] = len(met.reactions)
                else:
                    data[met_id] = [met.name, 0, len(met.reactions)]

        for met in metabolites:
            if periplasm and met.compartment == periplasm:
                met_id = str(met.id)
                if met_id in data:
                    data[met_id][1] = data[met_id][1] + len(met.reactions)
                else:
                    data[met_id] = [met.name, len(met.reactions), 0]

        df = DataFrame.from_dict(data, orient='index', columns=col)
        df.to_excel(writer, "Output")
        writer.save()
        writer.close()

    def topological_analysis(self, output_file_name):

        # open output file
        writer = pd.ExcelWriter(os.path.join(self.directory, output_file_name), engine='xlsxwriter')

        col = ["Inside", "Outside", "Gene", "Gene Rules"]

        reactions = self.reactions

        outside = 0
        inside = 0
        gene_rules = 0

        for reaction in reactions:

            if reaction.boundary:
                outside += 1
            else:
                inside += 1
                if reaction.gene_reaction_rule != '':
                    gene_rules += 1

        data = {"results": [inside, outside, len(self.genes), gene_rules]}

        df = DataFrame.from_dict(data, orient='index', columns=col)
        df.to_excel(writer, "Output")
        writer.save()
        writer.close()

    def revert_reaction(self, reaction_id):
        reaction = self.get_reaction(reaction_id)
        for met in reaction.metabolites:
            old_st = reaction.metabolites[met]
            reaction.add_metabolites({met: -old_st})
            reaction.add_metabolites({met: -old_st})

    def set_photoautotrophy(self, previous_carbon_source="EX_C00033__dra", photon_uptake=-855):
        self.reactions.EX_C00205__dra.bounds = (photon_uptake, 1000)
        self.reactions.get_by_id(previous_carbon_source).bounds = (0, 1000)
        self.reactions.e_Biomass__cytop.bounds = (0, 1000)
        self.reactions.e_Biomass_ht__cytop.bounds = (0, 0)
        self.objective = "e_Biomass__cytop"

    def set_heterotrophy(self, carbon_source="EX_C00033__dra", update_value=-10):
        self.reactions.EX_C00205__dra.bounds = (0, 1000)
        self.reactions.get_by_id(carbon_source).bounds = (update_value, 1000)
        self.reactions.e_Biomass_ht__cytop.bounds = (0, 1000)
        self.reactions.e_Biomass__cytop.bounds = (0, 0)
        self.objective = "e_Biomass_ht__cytop"

    def set_mixotrophy(self, carbon_source="EX_C00033__dra", update_value=-10, photon_uptake=-855):
        self.reactions.get_by_id(carbon_source).bounds = (update_value, 1000)
        self.reactions.EX_C00205__dra.bounds = (photon_uptake, 1000)
        self.reactions.e_Biomass__cytop.bounds = (0, 1000)
        self.reactions.e_Biomass_ht__cytop.bounds = (0, 0)
        self.objective = "e_Biomass__cytop"

    def set_prism_reaction(self, reaction_id):
        for reaction in self.reactions:
            if reaction.id.startswith("PRISM") and reaction.id != reaction_id:
                reaction.bounds = (0, 0)

    def adjust_biomass(self, new_value, suffix="v2", biomass_reaction_id=None):
        if not biomass_reaction_id:
            biomass_reaction_id = self.bio_reaction.id
        new_biomass = Reaction(id=f"e_Biomass_{suffix}__cytop", lower_bound=0, upper_bound=10000)
        if not self.biomass_composition:
            self.biomass_composition = biomass_reaction_id
        stoichiometries = update_st(copy.deepcopy(self.biomass_composition), new_value)
        new_biomass.add_metabolites({self.metabolites.get_by_id(key): value for key, value in stoichiometries.items()})
        new_biomass.add_metabolites(
            {key: value for key, value in self.reactions.get_by_id(biomass_reaction_id).metabolites.items() if
             not key.id.startswith("e_")})
        self.add_reactions([new_biomass])

    def determine_precursors(self, composition, units):
        if units == 'mol/mol':
            mol_mol = normalize(composition)
            mg__mol_mm = {}
            for key, value in mol_mol.items():
                mg__mol_mm[key] = convert_mmol_mol_to_g_molMM(value, self.metabolites.get_by_id(key).formula_weight)
            mmol__g_mm = convert_mg_molMM_to_mmolM_gMM(mol_mol, sum(mg__mol_mm.values()))
            print(mmol__g_mm)

    def adjust_precursors(self, reaction_id, composition, suffix="v2"):
        old_composition = {key: value for key, value in self.reactions.get_by_id(reaction_id).metabolites.items() if
                           value < 0}
        for key, value in composition.items():
            if self.metabolites.get_by_id(key) in old_composition:
                old_composition[self.metabolites.get_by_id(key)] = -value / self.metabolites.get_by_id(
                    key).formula_weight
        mol_mol = normalize(old_composition)
        mg__mol_mm = {}
        for key, value in mol_mol.items():
            mg__mol_mm[key] = convert_mmol_mol_to_g_molMM(value, key.formula_weight)
        mmol__g_mm = convert_mg_molMM_to_mmolM_gMM(mol_mol, sum(mg__mol_mm.values()))
        new_reaction = Reaction(id=f"{reaction_id.split('__')[0]}_{suffix}__{reaction_id.split('__')[1]}",
                                lower_bound=0, upper_bound=10000)
        for reactant, stoichiometry in mmol__g_mm.items():
            new_reaction.add_metabolites({reactant: -round(stoichiometry, 4)})
        new_reaction.add_metabolites({self.metabolites.e_Pigment__chlo: 1})
        self.add_reactions([new_reaction])
        self.reactions.get_by_id(reaction_id).bounds = (0, 0)

    def setup_condition(self, condition):
        for reaction in self.reactions:
            if reaction.id.startswith("e_Biomass") and f"trial{condition}" not in reaction.id:
                reaction.bounds = (0, 0)
            if reaction.id.startswith("e_Pigment") and f"trial{condition}" not in reaction.id:
                reaction.bounds = (0, 0)
        if condition == "default":
            self.reactions.e_Biomass__cytop.bounds = (0, 1000)
            self.reactions.e_Pigment__chlo.bounds = (0, 1000)
        else:
            self.objective = f"e_Biomass_trial{condition}__cytop"

    def infer_biomass_from_model(self, biomass_reaction_name="e_Biomass__cytop", biomass_met_id="e_Biomass__cytop"):
        """
        Infer biomass composition from biomass reaction
        :param biomass_reaction_name:
        :param biomass_met_id:
        :return:
        """
        biomass_reaction = self.reactions.get_by_id(biomass_reaction_name)
        macromolecules = biomass_reaction.reactants
        macromolecules = set(macromolecules) - {self.metabolites.get_by_id("C00001__cytop"),
                                                self.metabolites.get_by_id("C00002__cytop")}
        if biomass_met_id in [met.id for met in self.metabolites]:
            e_biomass = BiomassComponent(self.metabolites.get_by_id(biomass_met_id), 1, None)
        else:
            e_biomass = BiomassComponent(biomass_met_id, 1, None)
        self.biomass_components[biomass_reaction_name] = e_biomass
        for macromolecule in macromolecules:
            macromolecule_component = BiomassComponent(macromolecule.id, biomass_reaction.metabolites[macromolecule],
                                                       e_biomass)
            self.biomass_components[macromolecule.id] = macromolecule_component
            get_precursors(macromolecule_component, macromolecule, self)

    def get_reactions_pathways_map(self):
        pathway_map = {}
        for group in self.groups:
            for member in group.members:
                member_id = member.id
                if member_id in pathway_map:
                    pathway_map[member_id].append(group.name)
                else:
                    pathway_map[member_id] = [group.name]
        for reaction in self.reactions:
            if reaction.id not in pathway_map.keys():
                pathway_map[reaction.id] = []
        for key, value in pathway_map.items():
            pathway_map[key] = list(set(value))
        self.reactions_pathways_map = pathway_map

    def get_pathway_reactions_map(self):
        if not self.reactions_pathways_map:
            self.get_reactions_pathways_map()
        for key, value in self.reactions_pathways_map.items():
            for pathway in value:
                if pathway in self.pathway_reactions_map and key not in self.pathway_reactions_map[pathway]:
                    self.pathway_reactions_map[pathway].append(key)
                else:
                    self.pathway_reactions_map[pathway] = [key]

    def get_genes_pathways_map(self):
        res = {}
        if not self.reactions_pathways_map:
            self.get_reactions_pathways_map()
        for gene in self.genes:
            reactions = gene.reactions
            reactions_pathways = [self.reactions_pathways_map[reaction.id] for reaction in reactions]
            res[gene.id] = list(set([item for sublist in reactions_pathways for item in sublist]))
        self.genes_pathways_map = res
        return res

    def parse_genes(self):
        to_remove = []
        for gene in self.genes:
            if "_" in gene.id:
                other_gene = self.genes.get_by_id(gene.id.split("_")[0])
                other_gene.annotation = gene.annotation
                if not gene.reactions:
                    to_remove.append(gene)
        self.genes.remove(to_remove)

    def sample(self, method="achr", constraints=None):
        if constraints is None:
            constraints = {}
        for constraint in constraints.items():
            self.reactions.get_by_id(constraint[0]).bounds = constraint[1]
        if method == "achr":
            from cobra.sampling import ACHRSampler
            sampler = ACHRSampler(self, thinning=10, processes=6)
        elif method == "optgp":
            from cobra.sampling import OptGPSampler
            sampler = OptGPSampler(self, thinning=10, processes=6)
        else:
            raise ValueError("Method not supported")
        res = [s for s in sampler.batch(100, 10)]
        return res


def check_biomass(name, biomass):
    if biomass <= 0:
        print(name + "\t" * 3 + "No growth")
    elif biomass > 0:
        print(name + "\t" * 3 + "Growth" + "\t" * 3 + str(biomass))


def atp_m(model, m_reaction, mu, values=None):
    if values is None:
        values = {0.36: 0, 1: 0, 1.5: 0, 2: 0, 3: 0, 4: 0, 1.48: 0}
    m_reaction = model.get_reaction(m_reaction)
    original_bounds = m_reaction.bounds
    for key in values.keys():
        m_reaction.bounds = (key, key)
        values[key] = round(model.model.optimize().objective_value, 4)
    m_reaction.bounds = original_bounds
    x = np.array(list(values.keys())).reshape((-1, 1))
    y = np.array(list(values.values()))
    from sklearn.linear_model import LinearRegression
    regressor = LinearRegression()
    regressor.fit(x, y)
    print((mu - regressor.intercept_) / regressor.coef_)
    return values


def get_metabolite_compartment(model):
    c = 0
    e = 0
    unique_met = []
    for met in model.model.metabolites:
        if met.compartment == "C_00002":
            c += 1
            unique_met.append(met.name)
        elif met.compartment == "C_00001":
            e += 1
            unique_met.append(met.name)

    print("Cytoplasmic:", c, "\nExtracellular:", e, "\nUnique metabolites:", len(set(unique_met)))


def get_transport_number(model):
    tr = []
    for r in model.model.reactions:
        if r not in model.model.exchanges:
            cyt = False
            extr = False
            for met in r.metabolites:
                if met.compartment == "e":
                    extr = True
                if met.compartment == "c":
                    cyt = True
            if extr and cyt:
                tr.append(r)

    print("Number of transport reactions:", len(set(tr)))


def anaerobic(lr):
    lr.model.reactions.R00209__cytop.bounds = (0, 0)
    lr.model.exchanges.EX_C00007__dra.bounds = (0, 99999)
    lr.model.reactions.R00212__cytop.bounds = (0, 99999)
    lr.model.reactions.R00258__cytop.bounds = (-99999, 0)
    lr.model.reactions.R01827__cytop.bounds = (0, 99999)


def aerobic(lr):
    lr.model.exchanges.EX_C00007__dra.bounds = (-3.61, 999999)
    lr.model.reactions.R00209__cytop.bounds = (0, 99999)
    lr.model.reactions.R00212__cytop.bounds = (0, 0)
    lr.model.reactions.R01827__cytop.bounds = (0, 99999)


def define_medium(model, m, sheet=None, bigg=False):
    for ex in model.model.exchanges:
        ex.bounds = (0, 99999)
    if bigg:
        data = pd.read_excel("Media_bigg_updated.xlsx", sheet_name=sheet)
    else:
        data = pd.read_excel("Media_backup.xlsx", sheet_name=sheet)
    res = {}
    medium = data["M" + str(m)].dropna().to_list()
    for el in medium:
        res[el] = data["M" + str(m) + "_q"].loc[data["M" + str(m)] == el].values[0]
    for key in res:
        for ex in model.model.exchanges:
            if ex.id == "EX_" + key + "__dra" or ex.id == "EX_" + key + "_e":
                ex.bounds = (-res[key], 99999)

    try:
        if m == 3:
            model.model.reactions.R00754__cytop.bounds = (0, 0)
    except Exception as e:
        print(e)


def aerobic_v2(model):
    o = model.get_metabolite_by_name("Oxygen", "extr")
    for ex in model.model.exchanges:
        if o in ex.metabolites:
            ex.bounds = (-99999, 99999)
    return model


def compare_reacts(model1, model2):
    not_in_1 = []
    not_in_2 = []
    reacts_1 = model1.model.reactions
    reacts_2 = model2.model.reactions
    for r in reacts_1:
        if r not in reacts_2:
            not_in_2.append(r.id)
    for r in reacts_2:
        if r not in reacts_1:
            not_in_1.append(r.id)
    file = open("compare_models.csv", "w")
    if len(not_in_1) < len(not_in_2):
        for i in range(len(not_in_1)):
            file.write(str(not_in_1[i]) + "\t" + str(not_in_2[i]) + "\n")
        for i in range(len(not_in_1), len(not_in_2)):
            file.write("NAN" + "\t" + str(not_in_2[i]) + "\n")
    if len(not_in_1) >= len(not_in_2):
        for i in range(len(not_in_2)):
            file.write(str(not_in_1[i]) + "\t" + str(not_in_2[i]) + "\n")
        for i in range(len(not_in_2), len(not_in_1)):
            file.write(str(not_in_1[i]) + "\t" + "NAN" + "\n")
    file.close()


def define_medium_quercus(model):
    for exchange in model.exchanges:
        exchange.bounds = (0, 9999)
    model.exchanges.EX_C00205__dra.bounds = (-100, 999)
    model.exchanges.EX_C00011__dra.bounds = (-999, 999)
    model.exchanges.EX_C00001__dra.bounds = (-999, 999)
    model.exchanges.EX_C00014__dra.bounds = (-999, 999)
    model.exchanges.EX_C00080__dra.bounds = (-999, 999)
    model.exchanges.EX_C00009__dra.bounds = (-999, 999)
    model.exchanges.EX_C00305__dra.bounds = (-999, 999)
    model.exchanges.EX_C14818__dra.bounds = (-999, 999)
    model.exchanges.EX_C00059__dra.bounds = (-999, 999)
    model.exchanges.EX_C00007__dra.bounds = (-999, 999)


def react_without_gene(model):
    all_reactions = list(model.reactions)
    for r in all_reactions:
        if r in model.exchanges:
            all_reactions.remove(r)
    for gene in model.genes:
        for reaction in gene.reactions:
            if reaction in all_reactions:
                all_reactions.remove(reaction)
    for i in range(len(all_reactions)):
        all_reactions[i] = all_reactions[i].id
    print(all_reactions)


def photosynthesis(model):
    model_copy = model.copy()
    model_copy.reactions.R03140__chlo.bounds = (0, 0)
    model_copy.reactions.R00024__chlo.bounds = (0, 1000)
    model_copy.objective = "EX_C00205__dra"
    return model_copy


def photorespiration_v1(model, q=3):
    model_copy = model.model.copy()
    same_flux = model.model.problem.Constraint(
        model_copy.reactions.R00024__chlo.flux_expression - model_copy.reactions.R03140__chlo.flux_expression * q,
        lb=0,
        ub=0)
    model_copy.add_cons_vars(same_flux)
    model_copy.objective = "EX_C00205__dra"
    model_copy.reactions.e_Biomass_Leaf__cyto.bounds = (0.1, 0.1)
    model_copy.exchanges.EX_C00205__dra.lower_bound = -1000
    return model_copy


def photorespiration_v2(model, q=3):
    model_copy = model.model.copy()
    same_flux = model.model.problem.Constraint(
        model_copy.reactions.R00024__chlo.flux_expression - model_copy.reactions.R03140__chlo.flux_expression * q,
        lb=0,
        ub=0)
    model_copy.add_cons_vars(same_flux)
    model_copy.objective = "e_Biomass_Leaf__cyto"
    model_copy.exchanges.EX_C00205__dra.bounds = (-100, 999)
    return model_copy


def update_biomass(model, biomass_reaction, metabolites_to_remove):
    """
    NOT FINISHED
    """
    for met_id in metabolites_to_remove:
        metabolite = model.metabolites.get_by_id(met_id)
        if metabolite in model.reactions.get_by_id(biomass_reaction).metabolites:
            st = model.reactions.get_by_id(biomass_reaction).metabolites[metabolite]
            model.reactions.get_by_id(biomass_reaction).add_metabolites({model.metabolites.get_by_id(metabolite): -st})


def check_under_limit(reaction):
    balance = reaction.check_mass_balance()
    if balance:
        for key in balance:
            if round(balance[key], 5) != 0:
                return False
    return True


def check_balance(model, show_biomass_reactions=False):
    res = {}
    for reaction in model.reactions:
        if reaction.check_mass_balance() and "EX_" not in reaction.id and not check_under_limit(
                reaction) and "DM_" not in reaction.id:
            if str(reaction.id.split('__')[0]) + str(reaction.check_mass_balance()) not in res:
                if str(reaction.id).startswith("e_"):
                    if show_biomass_reactions:
                        res[str(reaction.id)] = reaction.check_mass_balance()
                else:
                    res[str(reaction.id)] = reaction.check_mass_balance()
    return res


def simulation_for_conditions(model, conditions_df, growth_rate_df, save_in_file=False, filename=None, objective=None):
    as_dict = conditions_df.to_dict(orient='index')
    growth_rate = growth_rate_df.to_dict(orient='index')
    complete_results = {}
    error_sum = 0
    values_for_plot = {}
    model.exchanges.EX_C00011__dra.bounds = (-1000, 1000)
    for index, condition in as_dict.items():
        model_copy = model.copy()
        for reaction in model_copy.reactions:
            if ("Biomass" in reaction.id and "EX_" not in reaction.id
                    and reaction.id != f"e_Biomass_trial{index}__cytop"):
                reaction.bounds = (0, 0)
        model_copy.reactions.get_by_id(f"e_Biomass_trial{index}__cytop").bounds = (0, 1000)
        if objective:
            [setattr(x, 'objective_coefficient', 0) for x in model.reactions if x.objective_coefficient != 0]
            model_copy.reactions.get_by_id(f"e_Biomass_trial{index}__cytop").objective_coefficient = 1
            for key, value in objective.items():
                model_copy.reactions.get_by_id(key).objective_coefficient = value
        else:
            model_copy.objective = f"e_Biomass_trial{index}__cytop"
        for met, lb in condition.items():
            lb = -lb if lb < 0 else lb
            model_copy.reactions.get_by_id("EX_" + met + "__dra").bounds = (round(-lb, 4), 1000)
        sol = model_copy.optimize()
        biomass = round(sol[f"e_Biomass_trial{index}__cytop"], 3)
        error_sum += abs(growth_rate[index]['growth_rate'] - biomass)
        complete_results[index] = sol
        values_for_plot[index] = (growth_rate[index]['growth_rate'], biomass)
    if save_in_file:
        write_simulation(complete_results, filename)
    return complete_results, values_for_plot, round(error_sum, 6)


def get_reactions_nadh_nadph(model):
    print("starting.....")
    for reaction in model.reactions:
        if model.metabolites.C00003__cytop in reaction.metabolites:
            for reaction2 in model.reactions:
                if reaction.genes == reaction2.genes and model.metabolites.C00005__cytop in reaction2.metabolites:
                    metabolites_1, metabolites_2 = [], []
                    for met in reaction.metabolites:
                        if met.id != "C00003__cytop" and met.id != "C00004__cytop":
                            metabolites_1.append(met.id)
                    for met in reaction2.metabolites:
                        if met.id != "C00005__cytop" and met.id != "C00006__cytop":
                            metabolites_2.append(met.id)
                    if metabolites_1 == metabolites_2:
                        print(reaction.id, reaction.name)
                        print(reaction2.id, reaction.name)
                        print("----------")


def add_reaction_string_to_dataframe(dataframe, model):
    dataframe['Reaction'] = np.nan
    for reaction in dataframe.index:
        reaction = model.reactions.get_by_id(reaction)
        reaction_as_string = ''
        for reactant in reaction.reactants:
            reaction_as_string += str(abs(reaction.get_coefficient(reactant.id))) + ' ' + reactant.name + ' + '
        reaction_as_string = reaction_as_string[0:len(reaction_as_string) - 3]
        if reaction.reversibility:
            reaction_as_string += ' <=> '
        else:
            reaction_as_string += ' => '
        for products in reaction.products:
            reaction_as_string += str(abs(reaction.get_coefficient(products.id))) + ' ' + products.name + ' + '
        if reaction_as_string.strip()[-1] == '+':
            reaction_as_string = reaction_as_string[0:len(reaction_as_string) - 3]

        dataframe['Reaction'].loc[dataframe.index == reaction] = reaction_as_string
    return dataframe


def count_reactions_by_compartment(model):
    compartments = {}
    for reaction in model.model.reactions:

        if len(reaction.compartments) > 1:
            membrane = check_transport(reaction)
            if membrane is not None:
                if membrane not in compartments.keys():
                    compartments[membrane] = 1
                else:
                    compartments[membrane] += 1
        else:
            reaction_compartment = list(reaction.compartments)[0]
            if str(reaction.id).endswith('extr_b') is False:
                compartment = model.model.compartments[reaction_compartment]
                if compartment not in compartments.keys():
                    compartments[compartment] = 1
                else:
                    compartments[compartment] += 1

    return compartments


def check_transport(reaction):
    d = {"['C_00004', 'C_00006']": 'Chloroplast_Membrane', "['C_00001', 'C_00004']": 'Plasma_Membrane',
         "['C_00002', 'C_00004']": 'Mitochondrial_Membrane', "['C_00003', 'C_00004']": 'Peroxisome_Membrane',
         "['C_00004', 'C_00008']": 'Vacuole_Membrane', "['C_00004', 'C_00005']": 'Golgi_Membrane',
         "['C_00004', 'C_00007']": 'Endoplasmic_Reticulum_Membrane',
         "['C_00006', 'C_00007']": 'Endoplasmic_Reticulum_Membrane'}

    compartments = list(reaction.compartments)
    compartments.sort()
    if str(compartments) not in d.keys():
        print(reaction)
    else:
        membrane = d[str(compartments)]
        return membrane
